import json
import os
import sys
import secrets
import io
import qrcode
import base64
from datetime import datetime, time, timedelta
from flask import Flask, render_template, request, redirect, url_for, flash, session, send_file, make_response, jsonify
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from functools import wraps
from werkzeug.security import generate_password_hash, check_password_hash
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import threading
import html
import requests

app = Flask(__name__)

# Production secret key - load from environment or generate persistent key
SECRET_KEY_FILE = 'secret_key.txt'
if os.path.exists(SECRET_KEY_FILE):
    with open(SECRET_KEY_FILE, 'r') as f:
        app.secret_key = f.read().strip()
else:
    # Generate and save secret key for production
    secret_key = secrets.token_hex(32)
    with open(SECRET_KEY_FILE, 'w') as f:
        f.write(secret_key)
    app.secret_key = secret_key

# Inisialisasi Rate Limiter
limiter = Limiter(
    get_remote_address,
    app=app,
    default_limits=["200 per day", "50 per hour"],
    storage_uri="memory://",
)

# File database
DATA_FILE = 'data.json'
ATTENDANCE_FILE = 'attendance.json'
USERS_FILE = 'users.json'
LEAVES_FILE = 'leaves.json'
SETTINGS_FILE = 'settings.json'

# File lock untuk mencegah race condition saat menulis JSON
file_locks = {
    DATA_FILE: threading.Lock(),
    ATTENDANCE_FILE: threading.Lock(),
    USERS_FILE: threading.Lock(),
    LEAVES_FILE: threading.Lock(),
    SETTINGS_FILE: threading.Lock(),
}

# Default settings
DEFAULT_SETTINGS = {
    'school_name': 'Sistem Absensi',
    'attendance_start_time': '07:00',
    'attendance_end_time': '09:00',
    'late_time': '07:30',
    'telegram_enabled': False,
    'telegram_bot_token': '',
    'telegram_admin_chat_id': '',
    'telegram_notify_checkin': True,
    'telegram_notify_late': True,
    'telegram_notify_absent': True,
    'telegram_notify_leave': True,
    'theme_color': 'blue',
    'enable_qr': True,
    'enable_leave': True,
    'enable_late_tracking': True,
}

# Generate CSRF token
def generate_csrf_token():
    if 'csrf_token' not in session:
        session['csrf_token'] = secrets.token_hex(16)
    return session['csrf_token']

app.jinja_env.globals['csrf_token'] = generate_csrf_token

# Decorator untuk login required
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'logged_in' not in session:
            flash('Silakan login terlebih dahulu', 'error')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

# Decorator untuk admin role required
def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'logged_in' not in session:
            flash('Silakan login terlebih dahulu', 'error')
            return redirect(url_for('login'))
        
        users = load_users()
        username = session.get('username')
        user_role = None
        
        for user in users['users']:
            if user['username'] == username:
                user_role = user.get('role', 'admin')
                break
        
        if user_role not in ['admin', 'super_admin']:
            flash('Anda tidak memiliki akses ke halaman ini', 'error')
            return redirect(url_for('dashboard'))
        
        return f(*args, **kwargs)
    return decorated_function

# Decorator untuk super admin required
def super_admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'logged_in' not in session:
            flash('Silakan login terlebih dahulu', 'error')
            return redirect(url_for('login'))
        
        users = load_users()
        username = session.get('username')
        user_role = None
        
        for user in users['users']:
            if user['username'] == username:
                user_role = user.get('role', 'admin')
                break
        
        if user_role != 'super_admin':
            flash('Hanya Super Admin yang dapat mengakses halaman ini', 'error')
            return redirect(url_for('dashboard'))
        
        return f(*args, **kwargs)
    return decorated_function

# Decorator untuk CSRF protection
def csrf_protect(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if request.method == 'POST':
            token = session.get('csrf_token')
            if not token or token != request.form.get('csrf_token'):
                flash('Token CSRF tidak valid atau kedaluwarsa', 'error')
                return redirect(url_for('index'))
            
            # Generate token baru setelah validasi
            session['csrf_token'] = secrets.token_hex(16)
        return f(*args, **kwargs)
    return decorated_function

# Fungsi untuk sanitasi input
def sanitize_input(text, max_length=100):
    """Sanitasi input untuk mencegah XSS dan membatasi panjang"""
    if not text:
        return ""
    text = html.escape(text.strip())
    return text[:max_length]

# ===== SETTINGS FUNCTIONS =====

def load_settings():
    """Load settings dengan error handling"""
    try:
        with file_locks[SETTINGS_FILE]:
            if os.path.exists(SETTINGS_FILE):
                with open(SETTINGS_FILE, 'r', encoding='utf-8') as f:
                    settings = json.load(f)
                    # Merge with defaults untuk backward compatibility
                    for key, value in DEFAULT_SETTINGS.items():
                        if key not in settings:
                            settings[key] = value
                    return settings
    except (json.JSONDecodeError, IOError) as e:
        print(f"Error loading settings: {e}")
    return DEFAULT_SETTINGS.copy()

def save_settings(settings):
    """Save settings dengan thread safety"""
    try:
        with file_locks[SETTINGS_FILE]:
            with open(SETTINGS_FILE, 'w', encoding='utf-8') as f:
                json.dump(settings, f, indent=4, ensure_ascii=False)
    except IOError as e:
        print(f"Error saving settings: {e}")
        flash('Gagal menyimpan pengaturan', 'error')

def get_attendance_times():
    """Get attendance times from settings"""
    settings = load_settings()
    try:
        start = datetime.strptime(settings['attendance_start_time'], '%H:%M').time()
        end = datetime.strptime(settings['attendance_end_time'], '%H:%M').time()
        late = datetime.strptime(settings['late_time'], '%H:%M').time()
        return start, end, late
    except:
        return time(7, 0), time(9, 0), time(7, 30)

# ===== TELEGRAM FUNCTIONS =====

def send_telegram_message(message):
    """Send message via Telegram bot"""
    settings = load_settings()
    
    if not settings.get('telegram_enabled'):
        return False
    
    bot_token = settings.get('telegram_bot_token', '').strip()
    chat_id = settings.get('telegram_admin_chat_id', '').strip()
    
    if not bot_token or not chat_id:
        return False
    
    try:
        url = f"https://api.telegram.org/bot{bot_token}/sendMessage"
        data = {
            'chat_id': chat_id,
            'text': message,
            'parse_mode': 'HTML'
        }
        response = requests.post(url, data=data, timeout=5)
        return response.status_code == 200
    except Exception as e:
        print(f"Telegram error: {e}")
        return False
        return False

def notify_checkin(student_name, student_id, status, time_str):
    """Notify admin about check-in via Telegram"""
    settings = load_settings()
    
    if status == 'Hadir' and not settings.get('telegram_notify_checkin'):
        return
    if status == 'Terlambat' and not settings.get('telegram_notify_late'):
        return
    
    emoji = "✅" if status == "Hadir" else "⚠️"
    message = f"{emoji} <b>Check-in Baru</b>\n\n"
    message += f"Nama: {student_name}\n"
    message += f"ID: {student_id}\n"
    message += f"Status: {status}\n"
    message += f"Waktu: {time_str}\n"
    message += f"Tanggal: {datetime.now().strftime('%d %B %Y')}"
    
    send_telegram_message(message)

def notify_absent_students():
    """Notify about absent students (called after attendance window closes)"""
    settings = load_settings()
    if not settings.get('telegram_notify_absent'):
        return
    
    data = load_data()
    attendance = load_attendance()
    today = datetime.now().strftime('%Y-%m-%d')
    
    present_today = attendance.get(today, {})
    total_students = len(data['students'])
    total_present = len(present_today)
    total_absent = total_students - total_present
    
    if total_absent > 0:
        message = f"📊 <b>Laporan Absensi Hari Ini</b>\n\n"
        message += f"Total Siswa: {total_students}\n"
        message += f"Hadir: {total_present}\n"
        message += f"Tidak Hadir: {total_absent}\n\n"
        
        # List absent students
        absent_list = []
        for student in data['students']:
            if student['id'] not in present_today:
                absent_list.append(f"- {student['name']} ({student['id']})")
        
        if absent_list:
            message += "<b>Siswa Tidak Hadir:</b>\n"
            message += "\n".join(absent_list[:10])  # Limit to 10
            if len(absent_list) > 10:
                message += f"\n...dan {len(absent_list) - 10} siswa lainnya"
        
        send_telegram_message(message)

def notify_leave_request(student_name, student_id, leave_type, reason):
    """Notify admin about new leave request"""
    settings = load_settings()
    if not settings.get('telegram_notify_leave'):
        return
    
    message = f"📝 <b>Pengajuan Izin Baru</b>\n\n"
    message += f"Nama: {student_name}\n"
    message += f"ID: {student_id}\n"
    message += f"Tipe: {leave_type.capitalize()}\n"
    message += f"Alasan: {reason[:100]}\n"
    message += f"Tanggal: {datetime.now().strftime('%d %B %Y %H:%M')}\n\n"
    message += "Silakan login ke admin panel untuk approve/reject."
    
    send_telegram_message(message)

# ===== DATA FUNCTIONS =====

def load_data():
    """Load data dengan error handling"""
    try:
        with file_locks[DATA_FILE]:
            if os.path.exists(DATA_FILE):
                with open(DATA_FILE, 'r', encoding='utf-8') as f:
                    return json.load(f)
    except (json.JSONDecodeError, IOError) as e:
        print(f"Error loading data: {e}")
    return {"students": []}

def save_data(data):
    """Save data dengan thread safety"""
    try:
        with file_locks[DATA_FILE]:
            with open(DATA_FILE, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=4, ensure_ascii=False)
    except IOError as e:
        print(f"Error saving data: {e}")
        flash('Gagal menyimpan data', 'error')

def load_attendance():
    """Load attendance dengan error handling"""
    try:
        with file_locks[ATTENDANCE_FILE]:
            if os.path.exists(ATTENDANCE_FILE):
                with open(ATTENDANCE_FILE, 'r', encoding='utf-8') as f:
                    return json.load(f)
    except (json.JSONDecodeError, IOError) as e:
        print(f"Error loading attendance: {e}")
    return {}

def save_attendance(attendance_data):
    """Save attendance dengan thread safety"""
    try:
        with file_locks[ATTENDANCE_FILE]:
            with open(ATTENDANCE_FILE, 'w', encoding='utf-8') as f:
                json.dump(attendance_data, f, indent=4, ensure_ascii=False)
    except IOError as e:
        print(f"Error saving attendance: {e}")
        flash('Gagal menyimpan data absensi', 'error')

def load_users():
    """Load users dengan error handling"""
    try:
        with file_locks[USERS_FILE]:
            if os.path.exists(USERS_FILE):
                with open(USERS_FILE, 'r', encoding='utf-8') as f:
                    return json.load(f)
    except (json.JSONDecodeError, IOError) as e:
        print(f"Error loading users: {e}")
    return {"users": []}

def save_users(users_data):
    """Save users dengan thread safety"""
    try:
        with file_locks[USERS_FILE]:
            with open(USERS_FILE, 'w', encoding='utf-8') as f:
                json.dump(users_data, f, indent=4, ensure_ascii=False)
    except IOError as e:
        print(f"Error saving users: {e}")
        flash('Gagal menyimpan data users', 'error')

def load_leaves():
    """Load leave requests dengan error handling"""
    try:
        with file_locks[LEAVES_FILE]:
            if os.path.exists(LEAVES_FILE):
                with open(LEAVES_FILE, 'r', encoding='utf-8') as f:
                    return json.load(f)
    except (json.JSONDecodeError, IOError) as e:
        print(f"Error loading leaves: {e}")
    return {"leaves": []}

def save_leaves(leaves_data):
    """Save leave requests dengan thread safety"""
    try:
        with file_locks[LEAVES_FILE]:
            with open(LEAVES_FILE, 'w', encoding='utf-8') as f:
                json.dump(leaves_data, f, indent=4, ensure_ascii=False)
    except IOError as e:
        print(f"Error saving leaves: {e}")
        flash('Gagal menyimpan data izin', 'error')

def get_user_role(username):
    """Get user role by username"""
    users = load_users()
    for user in users['users']:
        if user['username'] == username:
            return user.get('role', 'viewer')
    return None

def is_within_attendance_time():
    """Check apakah masih dalam waktu absensi"""
    start_time, end_time, _ = get_attendance_times()
    now = datetime.now().time()
    return start_time <= now <= end_time

def is_late():
    """Check apakah sudah terlambat"""
    _, _, late_time = get_attendance_times()
    now = datetime.now().time()
    return now > late_time

# ===== ROUTES =====

@app.route('/')
def index():
    settings = load_settings()
    start_time, end_time, late_time = get_attendance_times()
    
    return render_template('index.html', 
                           now=datetime.now(),
                           ATTENDANCE_START_TIME=start_time,
                           ATTENDANCE_END_TIME=end_time,
                           settings=settings)

@app.route('/qr_checkin')
def qr_checkin():
    """Halaman QR Code Check-in"""
    settings = load_settings()
    if not settings.get('enable_qr'):
        flash('Fitur QR Code tidak diaktifkan', 'error')
        return redirect(url_for('index'))
    
    start_time, end_time, _ = get_attendance_times()
    return render_template('qr_checkin.html',
                           now=datetime.now(),
                           ATTENDANCE_START_TIME=start_time,
                           ATTENDANCE_END_TIME=end_time,
                           settings=settings)

@app.route('/check_in', methods=['POST'])
@limiter.limit("10 per minute")
@csrf_protect
def check_in():
    student_id = sanitize_input(request.form.get('student_id'), max_length=50)
    
    if not student_id:
        flash('ID tidak boleh kosong', 'error')
        return redirect(url_for('index'))
    
    data = load_data()
    attendance = load_attendance()
    
    # Cek apakah ID ada di database
    student = None
    for s in data['students']:
        if s['id'] == student_id:
            student = s
            break
    
    if not student:
        flash('ID tidak terdaftar', 'error')
        return redirect(url_for('index'))
    
    # Cek apakah masih dalam waktu absensi
    if not is_within_attendance_time():
        flash('Waktu absensi telah berakhir', 'error')
        return redirect(url_for('index'))
    
    # Cek apakah sudah absen hari ini
    today = datetime.now().strftime('%Y-%m-%d')
    if today not in attendance:
        attendance[today] = {}
    
    if student_id in attendance[today]:
        flash('Anda sudah melakukan absensi hari ini', 'info')
        return redirect(url_for('index'))
    
    # Tentukan status (Hadir atau Terlambat)
    settings = load_settings()
    current_time = datetime.now().strftime('%H:%M:%S')
    
    if settings.get('enable_late_tracking') and is_late():
        status = 'Terlambat'
    else:
        status = 'Hadir'
    
    # Catat absensi
    attendance[today][student_id] = {
        'name': student['name'],
        'time': current_time,
        'status': status
    }
    save_attendance(attendance)
    
    # Send Telegram notification
    notify_checkin(student['name'], student_id, status, current_time)
    
    if status == 'Terlambat':
        flash(f'Absensi berhasil, {student["name"]} - Anda terlambat!', 'warning')
    else:
        flash(f'Absensi berhasil, selamat pagi {student["name"]}!', 'success')
    return redirect(url_for('index'))

@app.route('/leave_request', methods=['GET', 'POST'])
def leave_request():
    """Halaman pengajuan izin/sakit"""
    settings = load_settings()
    if not settings.get('enable_leave'):
        flash('Fitur pengajuan izin tidak diaktifkan', 'error')
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        student_id = sanitize_input(request.form.get('student_id'), max_length=50)
        reason = sanitize_input(request.form.get('reason'), max_length=500)
        leave_type = sanitize_input(request.form.get('leave_type'), max_length=20)
        
        if not student_id or not reason or not leave_type:
            flash('Semua field harus diisi', 'error')
            return redirect(url_for('leave_request'))
        
        if leave_type not in ['sakit', 'izin']:
            flash('Tipe izin tidak valid', 'error')
            return redirect(url_for('leave_request'))
        
        data = load_data()
        
        # Cek apakah ID ada di database
        student = None
        for s in data['students']:
            if s['id'] == student_id:
                student = s
                break
        
        if not student:
            flash('ID tidak terdaftar', 'error')
            return redirect(url_for('leave_request'))
        
        leaves = load_leaves()
        
        # Tambahkan leave request
        leave_id = f"LV{datetime.now().strftime('%Y%m%d%H%M%S')}"
        leaves['leaves'].append({
            'id': leave_id,
            'student_id': student_id,
            'student_name': student['name'],
            'date': datetime.now().strftime('%Y-%m-%d'),
            'time': datetime.now().strftime('%H:%M:%S'),
            'type': leave_type,
            'reason': reason,
            'status': 'pending'
        })
        
        save_leaves(leaves)
        
        # Send Telegram notification
        notify_leave_request(student['name'], student_id, leave_type, reason)
        
        flash(f'Pengajuan {leave_type} berhasil dikirim. Mohon tunggu persetujuan admin.', 'success')
        return redirect(url_for('leave_request'))
    
    return render_template('leave_request.html', settings=settings)

@app.route('/login', methods=['GET', 'POST'])
@limiter.limit("5 per minute")
def login():
    users = load_users()
    if not users.get('users'):
        return redirect(url_for('setup'))
    
    if request.method == 'POST':
        username = sanitize_input(request.form.get('username'), max_length=50)
        password = request.form.get('password', '')
        
        if not username or not password:
            flash('Username dan password harus diisi', 'error')
            return render_template('login.html')
        
        # Cek kredensial dengan password hash
        for user in users['users']:
            if user['username'] == username:
                if check_password_hash(user['password'], password):
                    session['logged_in'] = True
                    session['username'] = username
                    session['role'] = user.get('role', 'viewer')
                    flash('Login berhasil', 'success')
                    return redirect(url_for('dashboard'))
                else:
                    flash('Password salah', 'error')
                    return render_template('login.html')
        
        flash('Username tidak ditemukan', 'error')
    
    settings = load_settings()
    return render_template('login.html', settings=settings)

@app.route('/logout')
def logout():
    session.clear()
    flash('Anda telah logout', 'info')
    return redirect(url_for('index'))

@app.route('/dashboard')
@login_required
def dashboard():
    """Unified dashboard untuk semua role"""
    data = load_data()
    attendance = load_attendance()
    leaves = load_leaves()
    settings = load_settings()
    
    # Get user role
    username = session.get('username')
    user_role = get_user_role(username)
    
    # Hitung statistik absensi hari ini
    today = datetime.now().strftime('%Y-%m-%d')
    present_today = attendance.get(today, {})
    
    # Filter leave requests yang pending
    pending_leaves = [l for l in leaves['leaves'] if l['status'] == 'pending']
    
    # Hitung statistik
    total_students = len(data['students'])
    total_present = len(present_today)
    total_absent = total_students - total_present
    
    # Hitung yang terlambat hari ini
    late_today = sum(1 for record in present_today.values() if record.get('status') == 'Terlambat')
    
    return render_template('dashboard.html', 
                         students=data['students'], 
                         attendance=attendance,
                         present_today=present_today,
                         today=today,
                         now=datetime.now(),
                         pending_leaves=pending_leaves,
                         total_students=total_students,
                         total_present=total_present,
                         total_absent=total_absent,
                         late_today=late_today,
                         user_role=user_role,
                         settings=settings)

# Redirect old /admin to /dashboard for compatibility
@app.route('/admin')
@login_required
def admin():
    return redirect(url_for('dashboard'))

@app.route('/admin/add_student', methods=['POST'])
@admin_required
@csrf_protect
def add_student():
    student_id = sanitize_input(request.form.get('student_id'), max_length=50)
    name = sanitize_input(request.form.get('name'), max_length=100)
    class_name = sanitize_input(request.form.get('class_name', ''), max_length=50)
    
    if not student_id or not name:
        flash('ID dan nama tidak boleh kosong', 'error')
        return redirect(url_for('dashboard'))
    
    # Validasi ID hanya berisi angka dan huruf
    if not student_id.replace('-', '').replace('_', '').isalnum():
        flash('ID hanya boleh berisi huruf, angka, dash (-), dan underscore (_)', 'error')
        return redirect(url_for('dashboard'))
    
    data = load_data()
    
    # Cek apakah ID sudah ada
    for student in data['students']:
        if student['id'] == student_id:
            flash('ID sudah terdaftar', 'error')
            return redirect(url_for('dashboard'))
    
    # Tambahkan siswa baru
    data['students'].append({
        'id': student_id,
        'name': name,
        'class': class_name
    })
    
    save_data(data)
    flash('Siswa berhasil ditambahkan', 'success')
    return redirect(url_for('dashboard'))

@app.route('/admin/delete_student/<student_id>', methods=['POST'])
@admin_required
@csrf_protect
def delete_student(student_id):
    """Delete student dengan CSRF protection"""
    data = load_data()
    
    # Hapus siswa
    original_length = len(data['students'])
    data['students'] = [s for s in data['students'] if s['id'] != student_id]
    
    if len(data['students']) == original_length:
        flash('Siswa tidak ditemukan', 'error')
    else:
        save_data(data)
        flash('Siswa berhasil dihapus', 'success')
    
    return redirect(url_for('dashboard'))

@app.route('/admin/change_password', methods=['POST'])
@login_required
@csrf_protect
def change_password():
    current_password = request.form.get('current_password', '')
    new_password = request.form.get('new_password', '')
    confirm_password = request.form.get('confirm_password', '')
    
    if not current_password or not new_password or not confirm_password:
        flash('Semua field harus diisi', 'error')
        return redirect(url_for('dashboard'))
    
    if len(new_password) < 6:
        flash('Password baru minimal 6 karakter', 'error')
        return redirect(url_for('dashboard'))
    
    if new_password != confirm_password:
        flash('Password baru tidak cocok', 'error')
        return redirect(url_for('dashboard'))
    
    users = load_users()
    username = session.get('username')
    
    # Cek password saat ini dengan hash
    for user in users['users']:
        if user['username'] == username:
            if check_password_hash(user['password'], current_password):
                # Update password dengan hash
                user['password'] = generate_password_hash(new_password)
                save_users(users)
                flash('Password berhasil diubah', 'success')
                return redirect(url_for('dashboard'))
            else:
                flash('Password saat ini salah', 'error')
                return redirect(url_for('dashboard'))
    
    flash('User tidak ditemukan', 'error')
    return redirect(url_for('dashboard'))

@app.route('/admin/approve_leave/<leave_id>', methods=['POST'])
@admin_required
@csrf_protect
def approve_leave(leave_id):
    """Approve leave request"""
    leaves = load_leaves()
    
    for leave in leaves['leaves']:
        if leave['id'] == leave_id:
            leave['status'] = 'approved'
            leave['approved_by'] = session.get('username')
            leave['approved_at'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            # Tambahkan ke attendance sebagai izin/sakit
            attendance = load_attendance()
            date = leave['date']
            
            if date not in attendance:
                attendance[date] = {}
            
            attendance[date][leave['student_id']] = {
                'name': leave['student_name'],
                'time': leave['time'],
                'status': leave['type'].capitalize(),
                'reason': leave['reason']
            }
            
            save_attendance(attendance)
            save_leaves(leaves)
            flash(f"Pengajuan {leave['type']} telah disetujui", 'success')
            return redirect(url_for('dashboard'))
    
    flash('Pengajuan tidak ditemukan', 'error')
    return redirect(url_for('dashboard'))

@app.route('/admin/reject_leave/<leave_id>', methods=['POST'])
@admin_required
@csrf_protect
def reject_leave(leave_id):
    """Reject leave request"""
    leaves = load_leaves()
    
    for leave in leaves['leaves']:
        if leave['id'] == leave_id:
            leave['status'] = 'rejected'
            leave['rejected_by'] = session.get('username')
            leave['rejected_at'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            save_leaves(leaves)
            flash(f"Pengajuan {leave['type']} telah ditolak", 'info')
            return redirect(url_for('dashboard'))
    
    flash('Pengajuan tidak ditemukan', 'error')
    return redirect(url_for('dashboard'))

@app.route('/admin/export_csv')
@login_required
def export_csv():
    """Export attendance data to CSV"""
    attendance = load_attendance()
    
    # Buat CSV di memory
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Header
    writer.writerow(['Tanggal', 'ID Siswa', 'Nama', 'Waktu Absensi', 'Status'])
    
    # Data
    for date in sorted(attendance.keys(), reverse=True):
        for student_id, record in attendance[date].items():
            writer.writerow([
                date,
                student_id,
                record['name'],
                record['time'],
                record['status']
            ])
    
    # Buat response
    output.seek(0)
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'text/csv; charset=utf-8'
    response.headers['Content-Disposition'] = f'attachment; filename=attendance_{datetime.now().strftime("%Y%m%d")}.csv'
    
    return response

@app.route('/admin/export_excel')
@login_required
def export_excel():
    """Export attendance data to Excel dengan formatting"""
    attendance = load_attendance()
    
    # Buat workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Data Absensi"
    
    # Styling
    header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Header
    headers = ['Tanggal', 'ID Siswa', 'Nama', 'Waktu Absensi', 'Status']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Data
    row_num = 2
    for date in sorted(attendance.keys(), reverse=True):
        for student_id, record in attendance[date].items():
            ws.cell(row=row_num, column=1, value=date).border = border
            ws.cell(row=row_num, column=2, value=student_id).border = border
            ws.cell(row=row_num, column=3, value=record['name']).border = border
            ws.cell(row=row_num, column=4, value=record['time']).border = border
            
            status_cell = ws.cell(row=row_num, column=5, value=record['status'])
            status_cell.border = border
            
            # Color coding untuk status
            if record['status'] == 'Hadir':
                status_cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            elif record['status'] == 'Terlambat':
                status_cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
            elif record['status'] in ['Sakit', 'Izin']:
                status_cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
            
            row_num += 1
    
    # Auto-fit columns
    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = 20
    
    # Simpan ke BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Buat response
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=attendance_{datetime.now().strftime("%Y%m%d")}.xlsx'
    
    return response

@app.route('/admin/generate_qr/<student_id>')
@admin_required
def generate_qr(student_id):
    """Generate QR code untuk student"""
    data = load_data()
    settings = load_settings()
    
    # Cari student
    student = None
    for s in data['students']:
        if s['id'] == student_id:
            student = s
            break
    
    if not student:
        flash('Siswa tidak ditemukan', 'error')
        return redirect(url_for('dashboard'))
    
    # Generate QR code
    qr = qrcode.QRCode(version=1, box_size=10, border=5)
    qr.add_data(student_id)
    qr.make(fit=True)
    
    img = qr.make_image(fill_color="black", back_color="white")
    
    # Convert to base64
    buffer = io.BytesIO()
    img.save(buffer, format='PNG')
    buffer.seek(0)
    img_base64 = base64.b64encode(buffer.getvalue()).decode()
    
    return render_template('qr_display.html', 
                          student=student, 
                          qr_code=img_base64,
                          settings=settings)

# ===== ADMIN SETTINGS ROUTES =====

@app.route('/admin/settings', methods=['GET', 'POST'])
@super_admin_required
def admin_settings():
    """Admin settings panel"""
    if request.method == 'POST':
        settings = load_settings()
        
        # General settings
        settings['school_name'] = sanitize_input(request.form.get('school_name', ''), max_length=100)
        settings['attendance_start_time'] = sanitize_input(request.form.get('attendance_start_time', '07:00'), max_length=5)
        settings['attendance_end_time'] = sanitize_input(request.form.get('attendance_end_time', '09:00'), max_length=5)
        settings['late_time'] = sanitize_input(request.form.get('late_time', '07:30'), max_length=5)
        settings['theme_color'] = sanitize_input(request.form.get('theme_color', 'blue'), max_length=20)
        
        # Feature toggles
        settings['enable_qr'] = request.form.get('enable_qr') == 'on'
        settings['enable_leave'] = request.form.get('enable_leave') == 'on'
        settings['enable_late_tracking'] = request.form.get('enable_late_tracking') == 'on'
        
        # Telegram settings
        settings['telegram_enabled'] = request.form.get('telegram_enabled') == 'on'
        settings['telegram_bot_token'] = sanitize_input(request.form.get('telegram_bot_token', ''), max_length=200)
        settings['telegram_admin_chat_id'] = sanitize_input(request.form.get('telegram_admin_chat_id', ''), max_length=50)
        settings['telegram_notify_checkin'] = request.form.get('telegram_notify_checkin') == 'on'
        settings['telegram_notify_late'] = request.form.get('telegram_notify_late') == 'on'
        settings['telegram_notify_absent'] = request.form.get('telegram_notify_absent') == 'on'
        settings['telegram_notify_leave'] = request.form.get('telegram_notify_leave') == 'on'
        
        save_settings(settings)
        flash('Pengaturan berhasil disimpan', 'success')
        return redirect(url_for('admin_settings'))
    
    settings = load_settings()
    return render_template('admin_settings.html', settings=settings)

@app.route('/admin/test_telegram', methods=['POST'])
@super_admin_required
def test_telegram():
    """Test Telegram bot connection"""
    message = "🤖 <b>Test Koneksi Telegram Bot</b>\n\n"
    message += "Jika Anda menerima pesan ini, berarti bot Telegram telah terkonfigurasi dengan benar!\n\n"
    message += f"Waktu: {datetime.now().strftime('%d %B %Y %H:%M:%S')}"
    
    success = send_telegram_message(message)
    
    if success:
        return jsonify({'success': True, 'message': 'Pesan test berhasil dikirim! Cek Telegram Anda.'})
    else:
        return jsonify({'success': False, 'message': 'Gagal mengirim pesan. Periksa Bot Token dan Chat ID.'})

# ===== USER MANAGEMENT ROUTES =====

@app.route('/admin/users')
@super_admin_required
def manage_users():
    """User management page"""
    users = load_users()
    settings = load_settings()
    return render_template('manage_users.html', users=users['users'], settings=settings)

@app.route('/admin/users/add', methods=['POST'])
@super_admin_required
@csrf_protect
def add_user():
    """Add new user"""
    username = sanitize_input(request.form.get('username'), max_length=50)
    password = request.form.get('password', '')
    role = sanitize_input(request.form.get('role'), max_length=20)
    
    if not username or not password or not role:
        flash('Semua field harus diisi', 'error')
        return redirect(url_for('manage_users'))
    
    if role not in ['super_admin', 'admin', 'teacher', 'viewer']:
        flash('Role tidak valid', 'error')
        return redirect(url_for('manage_users'))
    
    if len(password) < 6:
        flash('Password minimal 6 karakter', 'error')
        return redirect(url_for('manage_users'))
    
    users = load_users()
    
    # Cek apakah username sudah ada
    for user in users['users']:
        if user['username'] == username:
            flash('Username sudah terdaftar', 'error')
            return redirect(url_for('manage_users'))
    
    # Tambahkan user baru
    users['users'].append({
        'username': username,
        'password': generate_password_hash(password),
        'role': role,
        'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    })
    
    save_users(users)
    flash(f'User {username} berhasil ditambahkan dengan role {role}', 'success')
    return redirect(url_for('manage_users'))

@app.route('/admin/users/delete/<username>', methods=['POST'])
@super_admin_required
@csrf_protect
def delete_user(username):
    """Delete user"""
    # Prevent deleting yourself
    if username == session.get('username'):
        flash('Anda tidak dapat menghapus akun Anda sendiri', 'error')
        return redirect(url_for('manage_users'))
    
    users = load_users()
    
    # Cek apakah ada super_admin lain
    super_admins = [u for u in users['users'] if u.get('role') == 'super_admin']
    if len(super_admins) == 1 and super_admins[0]['username'] == username:
        flash('Tidak dapat menghapus satu-satunya Super Admin', 'error')
        return redirect(url_for('manage_users'))
    
    # Hapus user
    original_length = len(users['users'])
    users['users'] = [u for u in users['users'] if u['username'] != username]
    
    if len(users['users']) == original_length:
        flash('User tidak ditemukan', 'error')
    else:
        save_users(users)
        flash(f'User {username} berhasil dihapus', 'success')
    
    return redirect(url_for('manage_users'))

@app.route('/admin/users/change_role/<username>', methods=['POST'])
@super_admin_required
@csrf_protect
def change_user_role(username):
    """Change user role"""
    new_role = sanitize_input(request.form.get('new_role'), max_length=20)
    
    if new_role not in ['super_admin', 'admin', 'teacher', 'viewer']:
        flash('Role tidak valid', 'error')
        return redirect(url_for('manage_users'))
    
    # Prevent changing your own role
    if username == session.get('username'):
        flash('Anda tidak dapat mengubah role Anda sendiri', 'error')
        return redirect(url_for('manage_users'))
    
    users = load_users()
    
    for user in users['users']:
        if user['username'] == username:
            old_role = user.get('role', 'viewer')
            user['role'] = new_role
            save_users(users)
            flash(f'Role {username} berhasil diubah dari {old_role} menjadi {new_role}', 'success')
            return redirect(url_for('manage_users'))
    
    flash('User tidak ditemukan', 'error')
    return redirect(url_for('manage_users'))

# Route untuk setup admin pertama kali
@app.route('/setup', methods=['GET', 'POST'])
def setup():
    users = load_users()
    if users['users']:
        flash('Setup sudah dilakukan sebelumnya', 'info')
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        username = sanitize_input(request.form.get('username'), max_length=50)
        password = request.form.get('password', '')
        confirm_password = request.form.get('confirm_password', '')
        
        if not username or not password:
            flash('Username dan password harus diisi', 'error')
            return render_template('setup.html')
        
        if len(password) < 6:
            flash('Password minimal 6 karakter', 'error')
            return render_template('setup.html')
        
        if password != confirm_password:
            flash('Password tidak cocok', 'error')
            return render_template('setup.html')
        
        # Buat user super admin pertama
        users['users'].append({
            'username': username,
            'password': generate_password_hash(password),
            'role': 'super_admin',
            'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        })
        
        save_users(users)
        
        # Initialize default settings
        if not os.path.exists(SETTINGS_FILE):
            save_settings(DEFAULT_SETTINGS)
        
        flash('Super Admin berhasil dibuat. Silakan login.', 'success')
        return redirect(url_for('login'))
    
    return render_template('setup.html')

if __name__ == '__main__':
    # Buat file data.json jika belum ada
    if not os.path.exists(DATA_FILE):
        with open(DATA_FILE, 'w', encoding='utf-8') as f:
            json.dump({"students": []}, f, indent=4)
    
    # Buat file attendance.json jika belum ada
    if not os.path.exists(ATTENDANCE_FILE):
        with open(ATTENDANCE_FILE, 'w', encoding='utf-8') as f:
            json.dump({}, f, indent=4)
    
    # Buat file users.json jika belum ada
    if not os.path.exists(USERS_FILE):
        with open(USERS_FILE, 'w', encoding='utf-8') as f:
            json.dump({"users": []}, f, indent=4)
    
    # Buat file leaves.json jika belum ada
    if not os.path.exists(LEAVES_FILE):
        with open(LEAVES_FILE, 'w', encoding='utf-8') as f:
            json.dump({"leaves": []}, f, indent=4)
    
    # Buat file settings.json jika belum ada
    if not os.path.exists(SETTINGS_FILE):
        with open(SETTINGS_FILE, 'w', encoding='utf-8') as f:
            json.dump(DEFAULT_SETTINGS, f, indent=4)
    
    # Production mode - use gunicorn instead
    print("For production, use: gunicorn -w 4 -b 0.0.0.0:5000 wsgi:application")
    print("For development, use: python app.py --dev")
    
    # Allow dev mode with flag
    if '--dev' in sys.argv:
        app.run(debug=True, host='0.0.0.0', port=5000)
    else:
        app.run(debug=False, host='0.0.0.0', port=5000)

